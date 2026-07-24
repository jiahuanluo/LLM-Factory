"""Vocab 构建：码值表 xlsx → cat_vocab.json.

cat_vocab.json 格式：
{
  "user": {"性别代码表": {"<UNK>": 0, "0": 1, "1": 2, ...}, ...},
  "account": {"<表名>": {...}, ...},
  ...
}

id 0 固定为 <UNK>（缺失/未见）。
"""
from __future__ import annotations

import json
from pathlib import Path

from .fields import (
    USER_CAT_FIELDS, ACCOUNT_CAT_FIELDS, QUERY_CAT_FIELDS,
    SUMMARY_TABLES, PAYSTATE_VOCAB, PAYSTATE_VOCAB_SIZE,
    PUBLIC_TYPE_VOCAB, PUBLIC_TYPE_VOCAB_SIZE,
)


def build_vocab_from_codetable(xlsx_path: str | Path) -> dict:
    """从 个人征信码值表.xlsx 构建每张表的 code → id 映射。"""
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("需要 openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # group by 第一列（关键字 / 码值表名）
    tables: dict[str, dict[str, int]] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name in ('关键字', '关键字 ',):
            continue
        code = row[1]
        if code is None:
            continue
        code = str(code).strip()
        tables.setdefault(name, {}).setdefault('<UNK>', 0)
        if code not in tables[name]:
            tables[name][code] = len(tables[name])

    return tables


def collect_used_tables() -> dict[str, list[str]]:
    """收集代码中真正用到的码值表名，按分支分组。"""
    used = {'user': [], 'summary': [], 'account': [], 'query': []}
    for path, table in USER_CAT_FIELDS:
        if table:
            used['user'].append(table)
    for _path, is_list, _nums, cats in SUMMARY_TABLES:
        for _f, t in cats:
            if t:
                used['summary'].append(t)
    for _path, table in ACCOUNT_CAT_FIELDS:
        if table:
            used['account'].append(table)
    for _path, table in QUERY_CAT_FIELDS:
        if table:
            used['query'].append(table)
    return used


def build_cat_vocab(xlsx_path: str | Path | None = None) -> dict:
    """构建完整 cat_vocab（含特殊 vocab 如 paystate、public_type）。"""
    if xlsx_path is None:
        # 默认路径
        p = Path('data/home-credit/个人征信/个人征信码值表.xlsx')
        if not p.exists():
            p = Path('data/pbc/个人征信码值表.xlsx')
        xlsx_path = p

    tables = build_vocab_from_codetable(xlsx_path)
    used = collect_used_tables()

    vocab: dict[str, dict[str, dict]] = {}
    for branch, table_names in used.items():
        vocab[branch] = {}
        for name in table_names:
            if name in tables:
                vocab[branch][name] = tables[name]
            else:
                # 码值表里没找到，用空表（只有 UNK）
                vocab[branch][name] = {'<UNK>': 0}

    # 特殊 vocab
    vocab['paystate'] = {'<all>': PAYSTATE_VOCAB}
    vocab['public_type'] = {'<all>': PUBLIC_TYPE_VOCAB}

    return vocab


def save_vocab(vocab: dict, out_path: str | Path):
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(vocab, f, ensure_ascii=False, indent=2)


def load_vocab(path: str | Path) -> dict:
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def get_vocab_size(branch: str, table: str, vocab: dict) -> int:
    """返回某分支某表 vocab 大小（含 UNK）。"""
    if branch in ('paystate', 'public_type'):
        return len(vocab[branch]['<all>'])
    return len(vocab.get(branch, {}).get(table, {'<UNK>': 0}))


def encode_value(branch: str, table: str, value, vocab: dict) -> int:
    """把码值编码成 id；空值/未知都返回 0 (<UNK>)。"""
    if value is None or value == '':
        return 0
    if branch in ('paystate', 'public_type'):
        return vocab[branch]['<all>'].get(value, 0)
    table_vocab = vocab.get(branch, {}).get(table, {})
    return table_vocab.get(str(value).strip(), 0)
