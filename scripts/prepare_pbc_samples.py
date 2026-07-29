"""基于 CrisPbc.json 模板造假数据集，用于本地端到端跑通。

输出（生产 JSONL 格式，每行 {biz_sno, pbcg2_json, label?}）：
  data/pbc/processed/train.jsonl
  data/pbc/processed/val.jsonl
  data/pbc/processed/cat_vocab.json

每条 sample 包含 pbc_struct 各模态 tensor（src/pbc_credit/sample_builder）。
文本分支（pbc_text / MLM）已拆分到 run_mlm.py 单独预训练，本脚本不再产出 text 字段。

用法：
  python scripts/prepare_pbc_samples.py --template data/home-credit/个人征信/CrisPbc.json --n 1000
"""
from __future__ import annotations

import argparse
import copy
import json
import random
import sys
from pathlib import Path

import numpy as np
import torch

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'src'))

from pbc_credit.sample_builder import build_sample, encode_sample
from pbc_credit.vocab import build_cat_vocab, save_vocab, load_vocab


def perturb_report(template: dict, idx: int, seed: int) -> dict:
    """对模板做随机扰动：改 reportsn + shuffle 账户 + 随机删除部分账户。"""
    rng = random.Random(seed + idx)
    r = copy.deepcopy(template)
    r['reportsn'] = f'mock_{idx:010d}_{rng.randint(10000, 99999)}'
    r['name'] = f'测试{idx:04d}'

    accs = r.get('accountInfos', [])
    rng.shuffle(accs)
    # 随机保留 2~8 个账户
    keep = rng.randint(min(2, len(accs)), min(8, len(accs)))
    r['accountInfos'] = accs[:keep]

    # 随机改部分金额数值（±10%）
    for a in r['accountInfos']:
        basic = a.get('accountBasic', {})
        for k in ('pd01aj01', 'pd01aj02', 'pd01aj03'):
            v = basic.get(k)
            if v and v != '':
                try:
                    fv = float(v)
                    basic[k] = str(int(fv * rng.uniform(0.9, 1.1)))
                except (TypeError, ValueError):
                    pass

    # shuffle queryRecords，随机保留
    qs = r.get('queryRecords', [])
    rng.shuffle(qs)
    r['queryRecords'] = qs[:rng.randint(min(3, len(qs)), min(20, len(qs)))]

    return r


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--template', default='data/home-credit/个人征信/CrisPbc.json')
    parser.add_argument('--out_dir', default='data/pbc/processed')
    parser.add_argument('--n', type=int, default=1000, help='总样本数')
    parser.add_argument('--pos_ratio', type=float, default=0.2)
    parser.add_argument('--val_ratio', type=float, default=0.15)
    parser.add_argument('--codetable', default='data/home-credit/个人征信/个人征信码值表.xlsx')
    parser.add_argument('--seed', type=int, default=42)
    args = parser.parse_args()

    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # === 1. 构建 vocab ===
    vocab_path = out_dir / 'cat_vocab.json'
    if not vocab_path.exists():
        print(f'=== 构建 cat_vocab from {args.codetable} ===')
        # 重定向路径让 build_cat_vocab 能找到码值表
        import openpyxl
        wb = openpyxl.load_workbook(args.codetable, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        tables = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            code = row[1]
            if code is None or name in ('关键字',):
                continue
            tables.setdefault(name, {}).setdefault('<UNK>', 0)
            code = str(code).strip()
            if code not in tables[name]:
                tables[name][code] = len(tables[name])
        # 用 build_cat_vocab 逻辑（传入已有 tables）
        from pbc_credit.vocab import collect_used_tables
        used = collect_used_tables()
        vocab = {
            branch: {name: tables.get(name, {'<UNK>': 0})
                     for name in names if name}
            for branch, names in used.items()
        }
        from pbc_credit.fields import PAYSTATE_VOCAB, PUBLIC_TYPE_VOCAB
        vocab['paystate'] = {'<all>': PAYSTATE_VOCAB}
        vocab['public_type'] = {'<all>': PUBLIC_TYPE_VOCAB}
        save_vocab(vocab, vocab_path)
        print(f'  saved → {vocab_path}')
    else:
        vocab = load_vocab(vocab_path)
        print(f'loaded vocab: {vocab_path}')

    # === 2. 读模板 + 造假 ===
    with open(args.template, encoding='utf-8') as f:
        template = json.load(f)

    n = args.n
    n_val = int(n * args.val_ratio)
    n_train = n - n_val

    splits = {
        'train': [(i, 1 if random.random() < args.pos_ratio else 0) for i in range(n_train)],
        'val': [(i + n_train, 1 if random.random() < args.pos_ratio else 0)
                for i in range(n_val)],
    }

    print(f'=== 造假 {n} 样本：train={n_train} val={n_val} ===')

    for split, items in splits.items():
        out_path = out_dir / f'{split}.jsonl'
        n_written = 0
        with open(out_path, 'w', encoding='utf-8') as f:
            for global_idx, label in items:
                r = perturb_report(template, global_idx, args.seed)
                sample = build_sample(r, vocab)
                encode_sample(sample, vocab)
                biz_sno = sample.pop('report_id', f'idx_{global_idx:010d}')
                sample.pop('target', None)  # label 走 JSONL 顶层，不进 pbcg2_json
                inner = {
                    k: (v.tolist() if hasattr(v, 'tolist') else v)
                    for k, v in sample.items()
                }
                record = {
                    'biz_sno': biz_sno,
                    'pbcg2_json': json.dumps(inner, ensure_ascii=False),
                }
                if label is not None:
                    record['label'] = int(label)
                f.write(json.dumps(record, ensure_ascii=False) + '\n')
                n_written += 1
        print(f'  wrote {n_written} → {out_path}')

    print(f'\n=== Done. Samples in {out_dir} ===')


if __name__ == '__main__':
    main()
