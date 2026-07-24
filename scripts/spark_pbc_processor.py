"""Spark PBC 处理作业 — 独立文件，无项目内 import 依赖。

把这个文件直接扔到 Spark 集群即可运行。依赖：
  - pyspark 3.x
  - Python 3.8+
  - openpyxl（构建 vocab 时用，运行时不需要）

============================================================
  生产部署流程
============================================================
1. 首次构建 cat_vocab.json（离线一次性）：
     python spark_pbc_processor.py build-vocab \
         --codetable 个人征信码值表.xlsx \
         --output /hdfs/path/cat_vocab.json

2. spark-submit 提交：
     spark-submit \
         --files /hdfs/path/cat_vocab.json \
         spark_pbc_processor.py run-spark \
         --input-table raw_pbc_reports \
         --input-column report_json \
         --output-table pbc_reports \
         --vocab-path cat_vocab.json

3. 或在 PySpark 交互环境直接用：
     from spark_pbc_processor import build_vocab_broadcast, parse_report_to_struct_json
     vocab_bc = build_vocab_broadcast(spark, '/hdfs/path/cat_vocab.json')
     parse_udf = spark.udf.register('parse_pbc', lambda s: parse_report_to_struct_json(s, vocab_bc.value))
     spark.sql("INSERT INTO pbc_reports SELECT report_id, pbc_text, parse_pbc(report_json) AS pbc_struct FROM raw_pbc_reports")
"""
from __future__ import annotations

import argparse
import datetime
import json
import math
import sys
from typing import Any


# ============================================================
# 一、字段定义（与 src/pbc_credit/fields.py 同步）
# ============================================================

# 账户类型（pd01ad01 取值）
ACCOUNT_TYPES = ['D1', 'R1', 'R2', 'R3', 'R4']

# 60 月还款状态字母表
PAYSTATE_VOCAB = {
    '<PAD>': 0, '<UNK>': 1,
    '#': 2, '*': 3, 'M': 4,
    '1': 5, '2': 6, '3': 7, '4': 8, '5': 9, '6': 10, '7': 11,
    'B': 12, 'C': 13, 'G': 14, 'D': 15, 'Z': 16, 'N': 17, 'A': 18, 'E': 19,
}

# User 分支（个人信息，固定维度）
# (json_path, code_table_name)；table=None 表示纯数值字段
USER_CAT_FIELDS = [
    ('personInfo.identity.pb01ad01', '性别代码表'),
    ('personInfo.identity.pb01ad02', '学历代码表'),
    ('personInfo.identity.pb01ad03', '学位代码表'),
    ('personInfo.identity.pb01ad04', '就业状况代码表'),
    ('personInfo.identity.pb01ad05', '世界各国和地区名称代码'),
    ('personInfo.marriage.pb020d01', '婚姻状况代码表'),
    ('personInfo.professionals.0.pb040d02', '单位性质代码表'),
    ('personInfo.professionals.0.pb040d03', '国民经济行业代码表'),
    ('personInfo.professionals.0.pb040d04', '职业代码表'),
    ('personInfo.professionals.0.pb040d05', '职务代码表'),
    ('personInfo.residences.0.pb030d01', '居住状况代码表'),
]

USER_NUMERIC_SPECS = [
    'age_years', 'num_mobiles', 'num_residences', 'num_professionals',
    'has_marriage', 'years_since_oldest_mobile', 'years_since_latest_mobile',
    'years_current_employer', 'has_email', 'num_identity_other_docs',
]

# Summary 分支（13 张概要子表聚合）
# (子表名, 是否 list, 数值字段列表, [(枚举字段, 码值表名)])
SUMMARY_TABLES = [
    ('tradeTips', True,
     ['pc02as01', 'pc02as03'],
     [('pc02ad01', '个人信贷交易提示业务类型代码表'),
      ('pc02ad02', '业务大类')]),
    ('recoveries', True,
     ['pc02bj01', 'pc02bj02'],
     [('pc02bd01', '个人被追偿汇总信息业务类型代码表')]),
    ('badDebit', False,
     ['pc02cj01'],
     [('pc02cs01', None)]),
    ('overdues', True,
     ['pc02dj01'],
     [('pc02dd01', '个人逾期（透支）汇总信息账户类型代码表'),
      ('pc02ds04', None)]),
    ('nonrevolvingLoan', False,
     ['pc02ej01', 'pc02ej02', 'pc02ej03'],
     [('pc02es01', None), ('pc02es02', None)]),
    ('revolvingCreditLoan', False,
     ['pc02fj01', 'pc02fj02', 'pc02fj03'],
     [('pc02fs01', None), ('pc02fs02', None)]),
    ('revolvingLoanAccount', False,
     ['pc02gj01', 'pc02gj02', 'pc02gj03'],
     [('pc02gs01', None), ('pc02gs02', None)]),
    ('loanCardAccount', False,
     ['pc02hj01', 'pc02hj02', 'pc02hj03', 'pc02hj04', 'pc02hj05'],
     [('pc02hs01', None), ('pc02hs02', None)]),
    ('standardLoancardAccount', False,
     ['pc02ij01', 'pc02ij02', 'pc02ij03', 'pc02ij04', 'pc02ij05'],
     [('pc02is01', None), ('pc02is02', None)]),
    ('relatedRepayDutys', True,
     ['pc02kj01', 'pc02kj02'],
     [('pc02kd01', '相关还款责任人类型代码表'),
      ('pc02kd02', '个人 借贷交易相关还款责任类型代码表')]),
    ('postpaySummary', False,
     ['pc030j01'],
     [('pc030d01', '后付费业务类型代码表')]),
    ('publics', True,
     ['pc040j01'],
     [('pc040d01', '公共信息类型代码表')]),
    ('querySummary', False,
     ['pc05bs01', 'pc05bs02', 'pc05bs03', 'pc05bs04',
      'pc05bs05', 'pc05bs06', 'pc05bs07', 'pc05bs08'],
     []),
]

# Account 分支
ACCOUNT_CAT_FIELDS = [
    ('pd01ad01', '个人借贷账户类型代码表'),
    ('pd01ad02', '个人借贷交易业务种类代码表'),
    ('pd01ad03', '个人借贷交易担保方式代码表'),
    ('pd01ad04', '币种代码表'),
    ('pd01ad06', '个人借贷交易还款频率代码表'),
    ('pd01ad07', '个人借贷交易担保方式代码表'),
]
ACCOUNT_NUMERIC_FIELDS = [
    'pd01ad05', 'pd01aj01', 'pd01aj02', 'pd01aj03',
    'pd01aj04', 'pd01as01', 'pd01bj01', 'pd01bj02',
]

# Query 分支
QUERY_CAT_FIELDS = [
    ('ph010d01', '机构类型代码'),
    ('ph010q03', '查询原因代码表'),
]
QUERY_NUMERIC_FIELDS = ['ph010r01_days_ago']  # 计算字段

# Public 分支
PUBLIC_TYPES = [
    ('pco_pf01', 'taxes'),
    ('pco_pf02', 'judgments'),
    ('pco_pf03', 'enforcement'),
    ('pco_pf04', 'penalties'),
    ('pco_pf05', 'low_income_relief'),
    ('pco_pf06', 'interest_arrears'),
    ('pco_pf07', 'professional_qual'),
    ('pco_pf08', 'awards'),
]
PUBLIC_TYPE_VOCAB = {name: i for i, (_node, name) in enumerate(PUBLIC_TYPES)}


# ============================================================
# 二、路径与日期工具
# ============================================================

def get_path(obj, path: str, default=None):
    """JSON 路径取值：'personInfo.identity.pb01ad01' 或 'list.0.field'。"""
    if not path:
        return default
    cur = obj
    for part in path.split('.'):
        if cur is None:
            return default
        if part.isdigit() and isinstance(cur, list):
            idx = int(part)
            cur = cur[idx] if idx < len(cur) else default
        elif isinstance(cur, dict):
            cur = cur.get(part, default)
        else:
            return default
    return cur if cur not in (None, '') else default


def parse_date(s):
    """支持 YYYY-MM-DD / YYYY-MM / YYYYMMDD / YYYYMMDDHHMMSS。"""
    if not s or not isinstance(s, str):
        return None
    s = s.strip()
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%Y-%m', '%Y%m%d', '%Y%m%d%H%M%S'):
        try:
            return datetime.datetime.strptime(s[:len(fmt)] if 'T' in fmt else s, fmt)
        except ValueError:
            continue
    return None


def years_since(dt, ref=None):
    if dt is None:
        return None
    if ref is None:
        ref = datetime.datetime.now()
    return (ref - dt).total_seconds() / (365.25 * 86400)


def days_since(dt, ref=None):
    if dt is None:
        return None
    if ref is None:
        ref = datetime.datetime.now()
    return (ref - dt).total_seconds() / 86400


def _report_ref_date(report: dict) -> datetime.datetime:
    """从报告 header 提参考日期（不硬编码，生产关键）。"""
    for path in ('tranDate', 'reportTime', 'header.request.tranDate'):
        dt = parse_date(get_path(report, path))
        if dt:
            return dt
    return datetime.datetime.now()


def _safe_float(v, default=None) -> float:
    if v is None or v == '':
        return default if default is not None else float('nan')
    try:
        return float(v)
    except (TypeError, ValueError):
        return default if default is not None else float('nan')


def _tame(v: float) -> float:
    """|v|>1000 走 log1p（压缩大金额）。"""
    if v is None or (isinstance(v, float) and v != v):
        return 0.0
    if abs(v) > 1000:
        return math.copysign(math.log1p(abs(v)), v)
    return float(v)


# ============================================================
# 三、Vocab（码值表 → id 映射）
# ============================================================

def build_vocab_from_codetable(xlsx_path: str) -> dict:
    """从 个人征信码值表.xlsx 构建 {码值表名: {code: id}}。"""
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("构建 vocab 需要 openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    tables = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        if name in ('关键字',):
            continue
        code = row[1]
        if code is None:
            continue
        code = str(code).strip()
        tables.setdefault(name, {}).setdefault('<UNK>', 0)
        if code not in tables[name]:
            tables[name][code] = len(tables[name])
    return tables


def _collect_used_tables() -> dict:
    """收集代码中用到的码值表名，按分支分组。"""
    used = {'user': [], 'summary': [], 'account': [], 'query': []}
    for _p, t in USER_CAT_FIELDS:
        if t:
            used['user'].append(t)
    for _n, _l, _nf, cats in SUMMARY_TABLES:
        for _f, t in cats:
            if t:
                used['summary'].append(t)
    for _f, t in ACCOUNT_CAT_FIELDS:
        if t:
            used['account'].append(t)
    for _f, t in QUERY_CAT_FIELDS:
        if t:
            used['query'].append(t)
    return used


def build_cat_vocab(xlsx_path: str) -> dict:
    """构建完整 cat_vocab（含 paystate / public_type 特殊 vocab）。"""
    tables = build_vocab_from_codetable(xlsx_path)
    used = _collect_used_tables()
    vocab = {}
    for branch, names in used.items():
        vocab[branch] = {}
        for name in names:
            vocab[branch][name] = tables.get(name, {'<UNK>': 0})
    vocab['paystate'] = {'<all>': PAYSTATE_VOCAB}
    vocab['public_type'] = {'<all>': PUBLIC_TYPE_VOCAB}
    return vocab


def load_vocab(path: str) -> dict:
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def save_vocab(vocab: dict, path: str):
    import os
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(vocab, f, ensure_ascii=False, indent=2)


def encode_value(branch: str, table: str, value, vocab: dict) -> int:
    """码值编码成 id；空值/未知都返回 0。"""
    if value is None or value == '':
        return 0
    if branch in ('paystate', 'public_type'):
        return vocab[branch]['<all>'].get(value, 0)
    table_vocab = vocab.get(branch, {}).get(table, {})
    return table_vocab.get(str(value).strip(), 0)


# ============================================================
# 四、Sample 构造（5 模态）
# ============================================================

def build_user(report: dict, ref: datetime.datetime) -> dict:
    """user 分支：10 numeric + 11 cat。"""
    identity = get_path(report, 'personInfo.identity', {}) or {}
    mobiles = get_path(report, 'personInfo.identity.mobiles', []) or []
    residences = get_path(report, 'personInfo.residences', []) or []
    professionals = get_path(report, 'personInfo.professionals', []) or []
    marriage = get_path(report, 'personInfo.marriage', {}) or {}
    identity_others = get_path(report, 'header.identityOthers', []) or []

    dob = parse_date(identity.get('pb01ar01'))
    age = years_since(dob, ref) if dob else None

    latest_mobile = earliest_mobile = None
    for m in mobiles:
        d = parse_date(m.get('pb01br01'))
        if d:
            if latest_mobile is None or d > latest_mobile:
                latest_mobile = d
            if earliest_mobile is None or d < earliest_mobile:
                earliest_mobile = d

    employer_year = None
    if professionals:
        y = professionals[0].get('pb040r01')
        if y:
            try:
                employer_year = datetime.datetime(int(y), 1, 1)
            except (ValueError, TypeError):
                pass

    numeric = [
        age if age is not None else 0.0,
        float(len(mobiles)),
        float(len(residences)),
        float(len(professionals)),
        1.0 if marriage else 0.0,
        years_since(earliest_mobile, ref) or 0.0,
        years_since(latest_mobile, ref) or 0.0,
        years_since(employer_year, ref) or 0.0,
        1.0 if identity.get('pb01aq01') else 0.0,
        float(len(identity_others)),
    ]
    # nan → 0
    numeric = [0.0 if (isinstance(x, float) and x != x) else x for x in numeric]

    cat_values = [get_path(report, path) for path, _t in USER_CAT_FIELDS]
    cat_mask = [0 if (v is None or v == '') else 1 for v in cat_values]
    return {'numeric': numeric, 'cat_values': cat_values, 'cat_mask': cat_mask}


def build_summary(report: dict) -> dict:
    """summary 分支：13 张概要子表聚合。"""
    sinfo = get_path(report, 'summaryInfo', {}) or {}
    nums, cats, cmask = [], [], []

    for name, is_list, num_fields, cat_fields in SUMMARY_TABLES:
        node = sinfo.get(name)
        if node is None:
            for _ in num_fields:
                nums.append(0.0)
            for _ in cat_fields:
                cats.append(None); cmask.append(0)
            continue

        if is_list:
            items = node if isinstance(node, list) else []
            nums.append(float(len(items)))
            for nf in num_fields:
                total = sum(_safe_float(it.get(nf), 0.0) or 0.0 for it in items)
                nums.append(_tame(total))
            for cf, _t in cat_fields:
                cats.append(items[0].get(cf) if items else None)
                cmask.append(1 if items else 0)
        else:
            for nf in num_fields:
                nums.append(_tame(_safe_float(node.get(nf))))
            for cf, _t in cat_fields:
                v = node.get(cf)
                cats.append(v)
                cmask.append(0 if v in (None, '') else 1)

    return {'numeric': nums, 'cat_values': cats, 'cat_mask': cmask}


def _parse_paystate_to_60(report_account: dict) -> list:
    """从 latest5year 或 latest24PayState 提取 60 月 id 序列。"""
    det = get_path(report_account, 'latest5year.latest5yearDetails')
    if det and isinstance(det, list) and len(det) > 0:
        states = []
        for row in det:
            s = row.get('pd01ed01', '')
            ch = str(s).strip()[:1].upper() if s else ''
            states.append(ch)
        states = states[-60:]
        states = ['<PAD>'] * (60 - len(states)) + states
        return [PAYSTATE_VOCAB.get(ch if ch != '<PAD>' else '<PAD>', 1) for ch in states]

    s24 = get_path(report_account, 'latest24PayState.latest24state')
    if s24 and isinstance(s24, str):
        chars = list(s24.upper())[-24:]
        chars = ['<PAD>'] * (24 - len(chars)) + chars
        chars = ['<PAD>'] * (60 - len(chars)) + chars
        return [PAYSTATE_VOCAB.get(c, 1) if c != '<PAD>' else 0 for c in chars]

    return [0] * 60


def build_accounts(report: dict) -> dict:
    """按 pd01ad01 分桶到 5 类账户。"""
    accs = get_path(report, 'accountInfos', []) or []
    by_type = {t: [] for t in ACCOUNT_TYPES}
    for a in accs:
        basic = a.get('accountBasic', {}) or {}
        t = basic.get('pd01ad01', '').strip()
        if t in by_type:
            by_type[t].append(a)

    result = {}
    for t, items in by_type.items():
        n = len(items)
        numeric = []
        cat_values = []
        cat_mask = []
        paystate = []
        for a in items:
            basic = a.get('accountBasic', {}) or {}
            latest = a.get('latestInfo', {}) or {}
            row_num = []
            for f in ACCOUNT_NUMERIC_FIELDS:
                v = basic.get(f) or latest.get(f)
                fv = _safe_float(v, 0.0) or 0.0
                if abs(fv) > 1000:
                    fv = math.copysign(math.log1p(abs(fv)), fv)
                row_num.append(fv)
            numeric.append(row_num)
            row_cat = []
            row_cmask = []
            for f, _t in ACCOUNT_CAT_FIELDS:
                v = basic.get(f)
                row_cat.append(v)
                row_cmask.append(0 if v in (None, '') else 1)
            cat_values.append(row_cat)
            cat_mask.append(row_cmask)
            paystate.append(_parse_paystate_to_60(a))
        result[t] = {
            'numeric': numeric, 'cat_values': cat_values,
            'cat_mask': cat_mask, 'paystate': paystate,
            'mask': [1] * n,
        }
    return result


def build_queries(report: dict, ref: datetime.datetime) -> dict:
    """query 分支：每条 [log1p(days_ago)] + [机构类型, 查询原因]。"""
    recs = get_path(report, 'queryRecords', []) or []
    numeric, cat_values, cat_mask = [], [], []
    for r in recs:
        d = days_since(parse_date(r.get('ph010r01')), ref) or 0.0
        if d != d:  # NaN
            d = 0.0
        numeric.append([math.log1p(max(0.0, d))])
        row_cat = []
        row_cmask = []
        for f, _t in QUERY_CAT_FIELDS:
            v = r.get(f)
            row_cat.append(v)
            row_cmask.append(0 if v in (None, '') else 1)
        cat_values.append(row_cat)
        cat_mask.append(row_cmask)
    return {'numeric': numeric, 'cat_values': cat_values, 'cat_mask': cat_mask}


def build_publics(report: dict, ref: datetime.datetime) -> dict:
    """public 分支：8 子表统一映射为 (amount_log1p, days_ago_log1p, type_id)。"""
    pinfo = get_path(report, 'publicInfo', {}) or {}
    numeric, cat_values, cmask = [], [], []
    for _node, type_name in PUBLIC_TYPES:
        key = _node.split('_')[-1]
        items = pinfo.get(key) or pinfo.get(key.upper())
        if not items:
            continue
        if isinstance(items, dict):
            items = [items]
        for it in items:
            amount = 0.0
            for k, v in it.items():
                if 'j01' in k.lower():
                    amount = _safe_float(v, 0.0) or 0.0
                    break
            days_ago = 0.0
            for k in it:
                if 'r01' in k.lower():
                    d = days_since(parse_date(it.get(k)), ref)
                    if d is None or d != d:
                        days_ago = 0.0
                    else:
                        days_ago = math.log1p(max(0.0, d))
                    break
            if abs(amount) > 1000:
                amount = math.copysign(math.log1p(abs(amount)), amount)
            numeric.append([amount, days_ago])
            cat_values.append([type_name])
            cmask.append([1])
    return {'numeric': numeric, 'cat_values': cat_values, 'cat_mask': cmask}


# ============================================================
# 五、主入口：单份 JSON → sample dict（已编码，纯 list）
# ============================================================

def build_sample(report: dict, vocab: dict) -> dict:
    """把 CrisPbc.json 解析为 sample dict（纯 list 格式，可直接 json.dumps）。

    生产端 Spark UDF 调这个；训练端 from_jsonable 反序列化即可。
    """
    ref = _report_ref_date(report)
    sample = {}

    # user
    u = build_user(report, ref)
    sample['user_numeric'] = u['numeric']
    sample['user_cat_ids'] = [
        encode_value('user', table, v, vocab)
        for v, (_p, table) in zip(u['cat_values'], USER_CAT_FIELDS)
    ]
    sample['user_cat_mask'] = u['cat_mask']

    # summary
    s = build_summary(report)
    sample['summary_numeric'] = s['numeric']
    summary_tables = [t for _n, _l, _nf, cf in SUMMARY_TABLES for _f, t in cf]
    sample['summary_cat_ids'] = [
        encode_value('summary', summary_tables[i] or '', v, vocab) if summary_tables[i] else 0
        for i, v in enumerate(s['cat_values'])
    ]
    sample['summary_cat_mask'] = s['cat_mask']

    # accounts
    acc_tables = [t for _f, t in ACCOUNT_CAT_FIELDS]
    accounts = build_accounts(report)
    for ty in ACCOUNT_TYPES:
        k = ty.lower()
        a = accounts[ty]
        sample[f'{k}_numeric'] = a['numeric']
        sample[f'{k}_cat_ids'] = [
            [encode_value('account', acc_tables[j] or '', row[j], vocab) if acc_tables[j] else 0
             for j in range(len(row))]
            for row in a['cat_values']
        ]
        sample[f'{k}_cat_mask'] = a['cat_mask']
        sample[f'{k}_paystate'] = a['paystate']
        sample[f'{k}_mask'] = a['mask']

    # queries
    q = build_queries(report, ref)
    q_tables = [t for _f, t in QUERY_CAT_FIELDS]
    sample['query_numeric'] = q['numeric']
    sample['query_cat_ids'] = [
        [encode_value('query', q_tables[j] or '', row[j], vocab) if q_tables[j] else 0
         for j in range(len(row))]
        for row in q['cat_values']
    ]
    sample['query_cat_mask'] = q['cat_mask']
    sample['query_mask'] = [1] * len(q['numeric'])

    # publics
    p = build_publics(report, ref)
    sample['public_numeric'] = p['numeric']
    sample['public_cat_ids'] = [
        [PUBLIC_TYPE_VOCAB.get(row[0], 0)] for row in p['cat_values']
    ]
    sample['public_cat_mask'] = p['cat_mask']
    sample['public_mask'] = [1] * len(p['numeric'])

    sample['report_id'] = report.get('reportsn', '')
    return sample


def parse_report_to_struct_json(report_json_str: str, vocab: dict) -> str:
    """Spark UDF 主函数：JSON 字符串 → pbc_struct JSON 字符串。"""
    try:
        report = json.loads(report_json_str)
    except (ValueError, TypeError):
        return None
    try:
        sample = build_sample(report, vocab)
        return json.dumps(sample, ensure_ascii=False)
    except Exception as e:
        # 生产环境容错：单条解析失败不要拖死整个 Spark job
        return json.dumps({'_error': str(e), 'report_id': report.get('reportsn', '')})


# ============================================================
# 六、Spark 集成
# ============================================================

def build_vocab_broadcast(spark, vocab_path: str):
    """Spark broadcast vocab（executor 共享一份，避免每条记录都加载）。"""
    vocab = load_vocab(vocab_path)
    return spark.sparkContext.broadcast(vocab)


def register_udfs(spark, vocab_bc):
    """注册 parse_pbc UDF 到 SparkSession。

    用法：
      spark.sql("SELECT report_id, parse_pbc(report_json) AS pbc_struct FROM raw")
    """
    def _udf(report_json_str):
        if report_json_str is None:
            return None
        return parse_report_to_struct_json(report_json_str, vocab_bc.value)
    spark.udf.register('parse_pbc', _udf)
    return _udf


def run_spark_job(spark, input_table: str, input_column: str,
                  output_table: str, vocab_path: str,
                  text_column: str = None):
    """完整 Spark 作业：读原始 JSON 列 → 产 pbc_struct（+ 可选 pbc_text）→ 写新表。

    Args:
        input_table: 原始表名（含 report_json 列）
        input_column: JSON 列名
        output_table: 目标表名（含 pbc_struct STRING 列）
        vocab_path: HDFS/本地 cat_vocab.json 路径
        text_column: 如果已经有 pbc_text 列，一起透传到输出表（默认 None）
    """
    vocab_bc = build_vocab_broadcast(spark, vocab_path)
    register_udfs(spark, vocab_bc)

    select_cols = ['report_id']
    if text_column:
        select_cols.append(text_column)
    select_cols.append(f'parse_pbc({input_column}) AS pbc_struct')

    sql = f"""
        INSERT INTO {output_table}
        SELECT {', '.join(select_cols)}
        FROM {input_table}
    """
    print(f'Running SQL:\n{sql}')
    spark.sql(sql)


# ============================================================
# 七、本地 CLI（测试 / 构 vocab / spark-submit 入口）
# ============================================================

def _cli():
    parser = argparse.ArgumentParser(description='PBC Spark 处理（独立文件）')
    sub = parser.add_subparsers(dest='cmd', required=True)

    # build-vocab: 离线构建 cat_vocab.json
    p1 = sub.add_parser('build-vocab', help='从码值表 xlsx 构建 cat_vocab.json')
    p1.add_argument('--codetable', required=True, help='个人征信码值表.xlsx 路径')
    p1.add_argument('--output', required=True, help='输出 cat_vocab.json 路径')

    # test-single: 本地测单条 JSON
    p2 = sub.add_parser('test-single', help='单条 JSON → pbc_struct（本地测试）')
    p2.add_argument('--input', required=True, help='CrisPbc.json 路径')
    p2.add_argument('--vocab', required=True, help='cat_vocab.json 路径')
    p2.add_argument('--output', help='输出路径（不填则 stdout 预览）')

    # run-spark: spark-submit 入口
    p3 = sub.add_parser('run-spark', help='spark-submit 入口')
    p3.add_argument('--input-table', required=True)
    p3.add_argument('--input-column', default='report_json')
    p3.add_argument('--output-table', required=True)
    p3.add_argument('--vocab-path', required=True)
    p3.add_argument('--text-column', default=None, help='已有 pbc_text 列名（透传）')

    args = parser.parse_args()

    if args.cmd == 'build-vocab':
        print(f'building vocab from {args.codetable}')
        vocab = build_cat_vocab(args.codetable)
        save_vocab(vocab, args.output)
        n_tables = sum(len(v) for v in vocab.values())
        print(f'saved {n_tables} tables → {args.output}')

    elif args.cmd == 'test-single':
        vocab = load_vocab(args.vocab)
        with open(args.input, encoding='utf-8') as f:
            report_json = f.read()
        struct_str = parse_report_to_struct_json(report_json, vocab)
        if args.output:
            with open(args.output, 'w', encoding='utf-8') as f:
                f.write(struct_str)
            print(f'wrote → {args.output}')
        else:
            print(struct_str[:3000])
            if len(struct_str) > 3000:
                print(f'... (total {len(struct_str)} chars)')

    elif args.cmd == 'run-spark':
        from pyspark.sql import SparkSession
        spark = SparkSession.builder.appName('pbc-struct-parse').getOrCreate()
        run_spark_job(
            spark=spark,
            input_table=args.input_table,
            input_column=args.input_column,
            output_table=args.output_table,
            vocab_path=args.vocab_path,
            text_column=args.text_column,
        )
        spark.stop()


if __name__ == '__main__':
    _cli()
