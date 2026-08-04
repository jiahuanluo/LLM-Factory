"""本地模拟 Spark UDF：raw JSON → pbc_struct JSONL。

对应生产 spark_parse_pbc.py 的 parse_report_to_struct_json UDF。
不做 scenario 扰动，不加 label —— 只是把 JSON 报文解析成特征 JSON 字符串。

输出（每行 {biz_sno, pbc_struct}）：
  data/pbc/processed/pbc_struct.jsonl

用法：
  python scripts/simulate_spark_parse.py --in_dir data/pbc/cris_json_split data/pbc/cris_json_diverse
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / 'src'))

from pbc_credit.sample_builder import build_sample, encode_sample, to_jsonable
from pbc_credit.vocab import build_cat_vocab, save_vocab, load_vocab


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--in_dir', nargs='+',
                        default=['data/pbc/cris_json_split', 'data/pbc/cris_json_diverse'],
                        help='生产报告目录（可多个）')
    parser.add_argument('--out', default='data/pbc/processed/pbc_struct.jsonl',
                        help='输出 JSONL（biz_sno + pbc_struct）')
    parser.add_argument('--vocab', default='data/pbc/processed/cat_vocab.json')
    parser.add_argument('--codetable', default='data/pbc/个人征信码值表.xlsx')
    args = parser.parse_args()

    # === 1. vocab ===
    vocab_path = Path(args.vocab)
    if vocab_path.exists():
        vocab = load_vocab(vocab_path)
        print(f'loaded vocab: {vocab_path}')
    else:
        print(f'=== 构建 cat_vocab from {args.codetable} ===')
        import openpyxl
        wb = openpyxl.load_workbook(args.codetable, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        tables = {}
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            code = row[1]
            if code is None or name in ('关键字',):
                continue
            tables.setdefault(name, {}).setdefault('<UNK>', 0)
            code = str(code).strip()
            if code not in tables[name]:
                tables[name][code] = len(tables[name])
        from pbc_credit.vocab import collect_used_tables
        used = collect_used_tables()
        from pbc_credit.fields import PAYSTATE_VOCAB, PUBLIC_TYPE_VOCAB, OBLIGATION_TYPE_VOCAB
        vocab = {
            branch: {name: tables.get(name, {'<UNK>': 0})
                     for name in names if name}
            for branch, names in used.items()
        }
        vocab['paystate'] = {'<all>': PAYSTATE_VOCAB}
        vocab['public_type'] = {'<all>': PUBLIC_TYPE_VOCAB}
        vocab['obligation_type'] = {'<all>': OBLIGATION_TYPE_VOCAB}
        save_vocab(vocab, vocab_path)
        print(f'  saved → {vocab_path}')

    # === 2. 读所有 JSON 报文 ===
    report_files = []
    for in_dir in args.in_dir:
        p = Path(in_dir)
        if not p.is_dir():
            continue
        report_files.extend(sorted(p.glob('*.json')))
    if not report_files:
        raise FileNotFoundError(f'no .json files in {args.in_dir}')
    print(f'=== 加载 {len(report_files)} 份 raw JSON 报文 ===')

    # === 3. parse 每份报文 → pbc_struct（与 Spark UDF 一致）===
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    n_ok = n_err = 0
    errors: list[tuple[str, str]] = []
    with open(out_path, 'w', encoding='utf-8') as f:
        for fp in report_files:
            try:
                with open(fp, encoding='utf-8') as fin:
                    report = json.load(fin)
                sample = build_sample(report, vocab)
                encode_sample(sample, vocab)
                biz_sno = sample.pop('report_id', fp.stem)
                sample.pop('target', None)
                pbc_struct = json.dumps(to_jsonable(sample), ensure_ascii=False)
                f.write(json.dumps({
                    'biz_sno': biz_sno,
                    'pbc_struct': pbc_struct,
                }, ensure_ascii=False) + '\n')
                n_ok += 1
            except Exception as e:
                n_err += 1
                if len(errors) < 5:
                    errors.append((fp.name, f'{type(e).__name__}: {str(e)[:80]}'))

    print(f'=== Done. {n_ok} OK / {n_err} err → {out_path} ===')
    if errors:
        for name, err in errors:
            print(f'  ERR {name}: {err}')


if __name__ == '__main__':
    main()
